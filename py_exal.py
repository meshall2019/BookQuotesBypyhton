from openpyxl import *


workbook = workbook = load_workbook(filename="/home/meshal/Desktop/book.xlsx")#path of the file
sheet = workbook.active
num=2
try:

  for i in sheet.rows:
  
      if sheet.cell(row=num,column=1).value==None and sheet.cell(row=num,column=2).value==None:
        
          sheet["A"+str(num)] =input("أدخل رقم الصفحة: ")
          sheet["B"+str(num)] =input("أدخل إقتباس الصفحة")
          workbook.save(filename="/home/meshal/Desktop/book.xlsx")
          print("done")
     
      else:
          num=1+num

except:
    print("Error")

